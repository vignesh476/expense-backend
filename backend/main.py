# ================= FastAPI Core =================
from fastapi import (
    FastAPI,
    APIRouter,
    Depends,
    HTTPException,
    Body,
    Request,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer

# ================= Pydantic =================
from pydantic import BaseModel, EmailStr

# ================= Auth / Security =================
from jose import jwt, JWTError
import bcrypt

# ================= Database =================
from pymongo import MongoClient
from bson import ObjectId

# ================= Date & Time =================
from datetime import datetime, timedelta

# ================= Excel Export =================
from openpyxl import Workbook
import io
import tempfile

# ================= Email =================
import resend
import smtplib
from email.message import EmailMessage

# ================= Environment / Utils =================
from dotenv import load_dotenv, find_dotenv
from pathlib import Path
import os
import math
import logging

# ================= Typing =================
from typing import List, Dict, Any, Optional


# Load env (Render ignores .env, local uses it)
load_dotenv(find_dotenv())

logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Expense Tracker API")

# ================= CONFIG =================

# Secrets / Tokens
RESEND_API_KEY = os.getenv("RESEND_API_KEY")
MONGO_URI = os.getenv("MONGO_URI")
JWT_SECRET = os.getenv("JWT_SECRET", "dev_change_me_jwt_secret")
JWT_REFRESH_SECRET = os.getenv("JWT_REFRESH_SECRET", "dev_change_me_jwt_refresh_secret")

ACCESS_EXP = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 15))
REFRESH_EXP = int(os.getenv("REFRESH_TOKEN_EXPIRE_DAYS", 7))
# Guest token expiry (short-lived by default)
GUEST_ACCESS_MIN = int(os.getenv("GUEST_ACCESS_MIN", 60))
GUEST_REFRESH_DAYS = int(os.getenv("GUEST_REFRESH_DAYS", 1))

# Frontend
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:5173")
FRONTEND_URL_2 = os.getenv("FRONTEND_URL_2","https://expense-gamma-six.vercel.app")
EMAIL_FROM = os.getenv(
    "EMAIL_FROM",
    "Expense<buggaramvignesh@gmail.com>"
)

# Validate critical env vars
if not MONGO_URI:
    raise RuntimeError("❌ MONGO_URI is not set")

if not RESEND_API_KEY:
    logging.warning("⚠️ RESEND_API_KEY not set – emails disabled")

resend.api_key = RESEND_API_KEY

# ================= DATABASE =================

client = MongoClient(MONGO_URI)
db = client.expense_app

users_col = db.users
tx_col = db.transactions
trips_col = db.trips

# ================= AUTH =================

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/login")

# ================= CORS =================

# Base allowed origins
allowed_origins = {
    FRONTEND_URL,
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:3000",
    "http://localhost:4173",
    "http://127.0.0.1:4173",
    "https://marvelous-pastelito-f03022.netlify.app",
    FRONTEND_URL_2
}

# Optional: allow multiple frontend URLs via ENV
# Example on Render:
# FRONTEND_URLS=https://app.com,https://admin.app.com
extra_origins = os.getenv("FRONTEND_URLS")
if extra_origins:
    allowed_origins.update(
        origin.strip() for origin in extra_origins.split(",")
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=list(allowed_origins),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ================= SCHEMAS =================
class RegisterSchema(BaseModel):
    email: EmailStr
    password: str

class LoginSchema(BaseModel):
    email: EmailStr
    password: str

class GuestLoginSchema(BaseModel):
    nickname: Optional[str] = None

# ================= PASSWORD =================
def hash_password(password: str) -> str:
    password = password[:72]
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password[:72].encode(), hashed.encode())

# ================= JWT =================
def create_token(data: dict, secret: str, exp: timedelta):
    payload = data.copy()
    payload["exp"] = datetime.utcnow() + exp
    return jwt.encode(payload, secret, algorithm="HS256")

def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        user = users_col.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(401, "User not found")
        return user
    except JWTError:
        raise HTTPException(401, "Invalid token")



# ================= AUTH =================
@app.post("/register")
def register(data: RegisterSchema):
    if users_col.find_one({"email": data.email}):
        raise HTTPException(400, "Email already exists")
    if len(data.password) < 6:
        raise HTTPException(400, "Password too short")

    users_col.insert_one({
        "email": data.email,
        "password": hash_password(data.password),
        "created": datetime.utcnow()
    })
    return {"ok": True}

@app.post("/login")
def login(data: LoginSchema):
    user = users_col.find_one({"email": data.email})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(401, "Invalid credentials")

    access = create_token(
        {"sub": str(user["_id"])},
        JWT_SECRET,
        timedelta(minutes=ACCESS_EXP)
    )

    refresh = create_token(
        {"sub": str(user["_id"])},
        JWT_REFRESH_SECRET,
        timedelta(days=REFRESH_EXP)
    )

    return {
        "access_token": access,
        "refresh_token": refresh
    }

@app.post("/guest-login")
def guest_login(payload: GuestLoginSchema = Body(default=None)):
    nickname = (payload.nickname if payload and payload.nickname else "Guest").strip() or "Guest"

    # Enforce simple nickname constraints and uniqueness among guests
    if len(nickname) < 2 or len(nickname) > 32:
        raise HTTPException(400, "Nickname must be 2-32 characters")

    # case-insensitive uniqueness for guest nicknames
    existing = users_col.find_one({
        "is_guest": True,
        "nickname": {"$regex": f"^{nickname}$", "$options": "i"}
    })
    if existing:
        raise HTTPException(409, "Nickname already taken. Choose a different one.")

    doc = {
        "email": None,
        "password": None,
        "is_guest": True,
        "nickname": nickname,
        "created": datetime.utcnow(),
        "guest_expires_at": datetime.utcnow() + timedelta(days=GUEST_REFRESH_DAYS),
    }
    res = users_col.insert_one(doc)

    access = create_token(
        {"sub": str(res.inserted_id), "guest": True},
        JWT_SECRET,
        timedelta(minutes=GUEST_ACCESS_MIN),
    )
    refresh = create_token(
        {"sub": str(res.inserted_id), "guest": True},
        JWT_REFRESH_SECRET,
        timedelta(days=GUEST_REFRESH_DAYS),
    )

    return {
        "access_token": access,
        "refresh_token": refresh,
        "guest": True,
        "user": {"id": str(res.inserted_id), "nickname": nickname},
    }

@app.post("/refresh")
def refresh_token(token: str = Body(...)):
    try:
       
        payload = jwt.decode(token, JWT_REFRESH_SECRET, algorithms=["HS256"])
        access = create_token(
            {"sub": payload["sub"]},
            JWT_SECRET,
            timedelta(minutes=ACCESS_EXP)
        )
        return {"access_token": access}
    except JWTError:
        raise HTTPException(401, "Invalid refresh token")

# ================= FORGOT PASSWORD =================
# @app.post("/forgot-password")
# def forgot(email: EmailStr = Body(...)):
     
#     user = users_col.find_one({"email": email})
#     if not user:
#         return {"ok": True}

#     token = create_token(
#         {"sub": str(user["_id"])},
#         JWT_SECRET,
#         timedelta(minutes=15)
#     )

#     link = f"{FRONTEND_URL}/reset-password?token={token}"
#     print(link)
 
#     try:
#         # resend.Emails.send({
#         #     "from": EMAIL_FROM,
#         #     "to": email,
#         #     "subject": "Reset Password",
#         #     "html": f"<a href='{link}'>Reset Password</a>"
#         # })
#         params: resend.Emails.SendParams = {
#         "from":"Acme <onboarding@resend.dev>",
#         "to": [email],
#         "subject": "Expenses",
#          "html": f"<a href='{link}'>Reset Password</a>",
#             }
#         email: resend.Emails.SendResponse = resend.Emails.send(params)
#         return email
        
#         print("✅ Email sent")
#     except Exception as e:
#         print("❌ Email error:", e)


#     return {"ok": email}


@app.post("/forgot-password")
async def forgot(request: Request, email: Optional[EmailStr] = Body(None, embed=True)):
    # Accept multiple input shapes: {"email": "..."}, raw JSON string, form data, or query param
    if not email:
        # Try JSON body
        try:
            data = await request.json()
            if isinstance(data, dict):
                email = data.get("email") or data.get("Email") or data.get("userEmail")
            elif isinstance(data, str):
                email = data
        except Exception:
            pass

        # Try form body
        if not email:
            try:
                form = await request.form()
                email = form.get("email") if form else None
            except Exception:
                pass

        # Try query param
        if not email:
            email = request.query_params.get("email")

    # Validate email using Pydantic
    try:
        class _EmailModel(BaseModel):
            email: EmailStr
        email = _EmailModel(email=email).email
    except Exception:
        raise HTTPException(422, "Invalid or missing email")

    user = users_col.find_one({"email": email})

    # Do not reveal whether email exists
    if not user:
        return {"ok": True}

    token = create_token(
        {"sub": str(user["_id"])},
        JWT_SECRET,
        timedelta(minutes=15),
    )

    link = f"{os.getenv('FRONTEND_URL')}/reset-password?token={token}"

    try:
        msg = EmailMessage()
        msg["From"] = os.getenv("EMAIL_FROM", 'Expense<buggaramvignesh@gmail.com>')
        msg["To"] = email
        msg["Subject"] = "Reset Your Password"

        msg.set_content(f"Reset your password using this link: {link}")

        msg.add_alternative(
            f"""
            <html>
              <body>
                <h3>Password Reset</h3>
                <p>Click the button below to reset your password:</p>
                <a href="{link}"
                   style="
                     display:inline-block;
                     padding:10px 16px;
                     background:#2563eb;
                     color:white;
                     text-decoration:none;
                     border-radius:5px;
                   ">
                   Reset Password
                </a>
                <p style="font-size:12px;color:#777">
                  Link expires in 15 minutes
                </p>
              </body>
            </html>
            """,
            subtype="html",
        )

        with smtplib.SMTP(
            os.getenv("BREVO_SMTP_HOST"),
            int(os.getenv("BREVO_SMTP_PORT", 587)),
        ) as server:
            server.starttls()
            server.login(
                os.getenv("BREVO_SMTP_USER"),
                os.getenv("BREVO_SMTP_PASS"),
            )
            server.send_message(msg)

        return {"ok": True}

    except Exception as e:
        print("❌ Email error:", e)
        raise HTTPException(500, "Failed to send reset email")



@app.post("/reset-password")
def reset_password(
    token: str = Body(...),
    password: str = Body(...)
):
    if len(password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        user_id = payload.get("sub")

        if not user_id:
            raise HTTPException(400, "Invalid token")

        user = users_col.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(400, "User not found")

        users_col.update_one(
            {"_id": ObjectId(user_id)},
            {
                "$set": {
                    "password": hash_password(password),
                    "password_changed_at": datetime.utcnow()
                }
            }
        )

        return {"ok": True, "message": "Password reset successful"}

    except jwt.ExpiredSignatureError:
        raise HTTPException(400, "Reset link expired")

    except JWTError:
        raise HTTPException(400, "Invalid reset link")

 
# ================= TRANSACTIONS =================
from datetime import datetime

@app.post("/transaction")
def create_transaction(amount: float, type: str, description: str,user=Depends(get_current_user), created_at: str = None):
    # parse created_at if provided, else use now
    if created_at:
        try:
            created_at_dt = datetime.fromisoformat(created_at)
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DDTHH:MM"}
    else:
        created_at_dt = datetime.utcnow()  # default to now

    tx = {
        "amount": amount,
        "type": type,
        "description": description,
        "user_id": user["_id"],
        "created_at": created_at_dt
    }

    result = tx_col.insert_one(tx)
    return {"status": "success", "transaction": {
        "id": str(result.inserted_id),
        "amount": tx["amount"],
        "type": tx["type"],
        "description": tx["description"],
        "user_id": str(tx["user_id"]),
        "created_at": tx["created_at"].isoformat()
    }}




@app.get("/transactions")
def get_transactions(user=Depends(get_current_user)):
    txs = list(
        tx_col.find({"user_id": user["_id"]}).sort("created_at", -1)
    )

    return [
        {
            "id": str(t["_id"]),
            "amount": t["amount"],
            "type": t["type"],
            "description": t.get("description", ""),
            "created_at": t["created_at"],
        }
        for t in txs
    ]



@app.put("/transaction/{tx_id}")
def update_transaction(
    tx_id: str,
    amount: float,
    type: str,
    description: str,
    user=Depends(get_current_user),
):
    result = tx_col.update_one(
        {"_id": ObjectId(tx_id), "user_id": user["_id"]},
        {"$set": {
            "amount": amount,
            "type": type,
            "description": description,
        }}
    )

    if result.matched_count == 0:
        raise HTTPException(404, "Transaction not found")

    return {"ok": True}


@app.delete("/transaction/{tx_id}")
def delete_transaction(tx_id: str, user=Depends(get_current_user)):
    result = tx_col.delete_one(
        {"_id": ObjectId(tx_id), "user_id": user["_id"]}
    )

    if result.deleted_count == 0:
        raise HTTPException(404, "Transaction not found")

    return {"ok": True}


# ================= SUMMARY =================
def build_summary(user_id, monthly=False):
    txs = list(tx_col.find({"user_id": user_id}))
    today = datetime.utcnow()
    today_date = today.date()

    def calc(ts, t):
        return sum(x["amount"] for x in ts if x["type"] == t)

    if monthly:
        # Get transactions from current month
        month_txs = [t for t in txs if t["created_at"].year == today.year and t["created_at"].month == today.month]
        return {
            "period": f"{today.strftime('%B %Y')}",
            "income": calc(month_txs, "income"),
            "expense": calc(month_txs, "expense"),
            "entries": month_txs
        }
    else:
        # Get today's transactions
        today_txs = [t for t in txs if t["created_at"].date() == today_date]
        return {
            "today": {
                "income": calc(today_txs, "income"),
                "expense": calc(today_txs, "expense")
            },
            "entries": today_txs
        }

# ================= EXCEL =================
def create_excel(summary, monthly=False):
    wb = Workbook()
    ws = wb.active
    
    if monthly:
        ws.append(["Monthly Report - " + summary.get("period", "")])
        ws.append([])
        ws.append(["Income", summary["income"]])
        ws.append(["Expense", summary["expense"]])
        ws.append(["Net", summary["income"] - summary["expense"]])
        ws.append([])
        ws.append(["Date", "Type", "Amount", "Description"])
        for entry in summary["entries"]:
            ws.append([
                entry["created_at"].strftime("%Y-%m-%d %H:%M"),
                entry["type"],
                entry["amount"],
                entry.get("description", "")
            ])
    else:
        ws.append(["Income", "Expense"])
        ws.append([summary["today"]["income"], summary["today"]["expense"]])
    
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return tmp.name

@app.get("/download-excel")
def download(monthly: bool = False, user=Depends(get_current_user)):
    path = create_excel(build_summary(user["_id"], monthly=monthly), monthly=monthly)
    filename = "monthly-summary.xlsx" if monthly else "daily-summary.xlsx"
    return FileResponse(path, filename=filename)

# @app.post("/send-summary")
# def send_summary(monthly: bool = False, user=Depends(get_current_user)):
#     try:
#         path = create_excel(build_summary(user["_id"], monthly=monthly), monthly=monthly)
#         import base64

#         with open(path, "rb") as f:
#             encoded = base64.b64encode(f.read()).decode()

#         report_type = "Monthly" if monthly else "Daily"
#         params: resend.Emails.SendParams = {
#             "from": "Expense App <onboarding@resend.dev>",
#             "to": user["email"],
#             "subject": f"Your {report_type} Expense Summary",
#             "html": f"<p>Please find attached your {report_type.lower()} expense summary.</p>",
#             "attachments": [
#                 {
#                     "filename": f"expense-summary-{report_type.lower()}.xlsx",
#                     "content": encoded,
#                 }
#             ],
#         }
#         email: resend.Emails.SendResponse = resend.Emails.send(params)

#         if path and os.path.exists(path):
#             os.remove(path)
#         return {"ok": True, "message": f"{report_type} summary sent successfully"}
#     except Exception as e:
#         logging.error(f"Error sending summary: {str(e)}")
#         raise HTTPException(500, f"Failed to send summary: {str(e)}")

@app.post("/send-summary")
def send_summary(monthly: bool = False, user=Depends(get_current_user)):
    report_type = "Monthly" if monthly else "Daily"
    path = None

    try:
        # 1️⃣ Try generating summary
        summary = build_summary(user["_id"], monthly=monthly)
        path = create_excel(summary, monthly=monthly)

        subject = f"Your {report_type} Expense Summary"
        text_body = f"Please find attached your {report_type.lower()} expense summary."
        html_body = f"""
        <html>
          <body>
            <h3>{report_type} Expense Summary</h3>
            <p>Please find attached your {report_type.lower()} expense summary.</p>
            <hr />
            <p style="font-size:12px;color:#777">
              Sent by Expense Tracker App
            </p>
          </body>
        </html>
        """

    except Exception as e:
        logging.error(f"Summary generation failed: {e}")

        # ⚠️ FALLBACK MESSAGE
        subject = f"{report_type} Expense Summary (Processing)"
        text_body = (
            "We are currently preparing your expense summary.\n\n"
            "You will receive the detailed report shortly."
        )
        html_body = """
        <html>
          <body>
            <h3>Expense Summary</h3>
            <p>We are currently preparing your expense summary.</p>
            <p>You will receive the detailed report shortly.</p>
            <hr />
            <p style="font-size:12px;color:#777">
              Expense Tracker Team
            </p>
          </body>
        </html>
        """

    try:
        # 2️⃣ Build Email
        msg = EmailMessage()
        msg["From"] = os.getenv("EMAIL_FROM",'Expense<buggaramvignesh@gmail.com>')
        msg["To"] = user["email"]
        msg["Subject"] = subject

        msg.set_content(text_body)
        msg.add_alternative(html_body, subtype="html")

        # 3️⃣ Attach Excel if available
        if path and os.path.exists(path):
            with open(path, "rb") as f:
                msg.add_attachment(
                    f.read(),
                    maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=f"expense-summary-{report_type.lower()}.xlsx",
                )

        # 4️⃣ SMTP SEND (CORRECT ORDER)
        with smtplib.SMTP(
            os.getenv("BREVO_SMTP_HOST1",'smtp-relay.brevo.com'),
            int(os.getenv("BREVO_SMTP_PORT1",587)),
        ) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(
                os.getenv("BREVO_SMTP_USER"),
                os.getenv("BREVO_SMTP_PASS"),
            )
            server.send_message(msg)

        return {"ok": True, "message": "Email sent successfully"}

    except Exception as e:
        logging.error(f"SMTP send failed: {e}")
        raise HTTPException(
            status_code=500,
            detail="Email service temporarily unavailable",
        )

    finally:
        if path and os.path.exists(path):
            os.remove(path)



# @app.post("/send-emails")
# def send_mail() -> Dict:
#     params: resend.Emails.SendParams = {
#         "from":"Acme <onboarding@resend.dev>",
#         "to": [''],
#         "subject": "Expenses",
#         "html": "<strong>it works!</strong>",
#     }
#     email: resend.Emails.SendResponse = resend.Emails.send(params)
#     return email
# import smtplib
# from email.mime.text import MIMEText
# from email.mime.multipart import MIMEMultipart
# import os

# @app.post("/send-email1")
# def send_summary_email(to_email, html_content):
#     msg = MIMEMultipart()
#     msg["From"] = os.getenv("SENDER_EMAIL",'')
#     msg["To"] = to_email
#     msg["Subject"] = "Your Expense Summary"

#     msg.attach(MIMEText(html_content, "html"))

#     with smtplib.SMTP(os.getenv("BREVO_SMTP_HOST",''), int(os.getenv("BREVO_SMTP_PORT",'587'))) as server:
#         server.starttls()
#         server.login(
#             os.getenv("BREVO_SMTP_USER",''),
#             os.getenv("BREVO_SMTP_PASS",'')
#         )
#         server.send_message(msg)



# Import from main module
# from .main import db, get_current_user

# trips_col = db.trips
# router = APIRouter(prefix="/trips", tags=["trips"])


# Pydantic request bodies
class TripCreate(BaseModel):
    trip_name: str
    start_date: str = None  # ISO date string (YYYY-MM-DD) optional
    end_date: str = None


class ParticipantCreate(BaseModel):
    name: str


class TripExpenseCreate(BaseModel):
    paid_by: str
    amount: float
    description: str = ""
    created_at: str = None  # optional ISO datetime string


# Helpers
def to_objid(id_str: str):
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(400, "Invalid id")


from datetime import datetime, time

def parse_iso_date(value: str):
    if not value:
        return None
    d = datetime.fromisoformat(value).date()
    return datetime.combine(d, time.min)  # 00:00:00



def parse_iso_datetime(s: str):
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def serialize_trip(trip: dict) -> dict:
    """Convert Mongo types (ObjectId, datetimes) to JSON-serializable values."""
    if not trip:
        return trip
    t = dict(trip)
    # _id -> id
    if "_id" in t:
        try:
            t["id"] = str(t.pop("_id"))
        except Exception:
            t.pop("_id", None)
    # user_id
    if "user_id" in t:
        try:
            t["user_id"] = str(t["user_id"])
        except Exception:
            pass
    # datetimes to isoformat
    for k in ("created", "start_date", "end_date"):
        if k in t and hasattr(t[k], "isoformat"):
            try:
                t[k] = t[k].isoformat()
            except Exception:
                pass
    # expenses
    exs = []
    for e in t.get("expenses", []) or []:
        ee = dict(e)
        if "created_at" in ee and hasattr(ee["created_at"], "isoformat"):
            try:
                ee["created_at"] = ee["created_at"].isoformat()
            except Exception:
                pass
        exs.append(ee)
    t["expenses"] = exs
    return t


# ========== Endpoints ==========

@app.post("/trips")
def create_trip(payload: TripCreate, user=Depends(get_current_user)):
    doc = {
        "user_id": user["_id"],
        "trip_name": payload.trip_name,
        "start_date": parse_iso_date(payload.start_date) if payload.start_date else None,
        "end_date": parse_iso_date(payload.end_date) if payload.end_date else None,
        "participants": [],  # list of {"name": "..."}
        "expenses": [],  # list of {"paid_by","amount","description","created_at"}
        "created": datetime.utcnow(),
    }
    res = trips_col.insert_one(doc)
    created = {**doc, "_id": res.inserted_id}
    return {"ok": True, "id": str(res.inserted_id), "trip": serialize_trip(created)}


@app.get("/trips")
def get_trips(user=Depends(get_current_user)):
    rows = list(trips_col.find({"user_id": user["_id"]}, {}))
    return [serialize_trip(r) for r in rows]
# @router.get("/trips")
# async def list_trips(user=Depends(get_current_user)):
#     trip_col = db["trips"]
#     trips = list(trip_col.find(
#         {"user_id": user["_id"]},  # ← Compare ObjectId to ObjectId
#         {}))  # Don't exclude _id; convert it to string in response
    
    # Convert ObjectId fields to strings for JSON serialization
    for trip in trips:
        trip["_id"] = str(trip["_id"])
        if "user_id" in trip and hasattr(trip["user_id"], "__str__"):
            trip["user_id"] = str(trip["user_id"])
    
    return trips

@app.get("/trips/{trip_id}")
def get_trip(trip_id: str, user=Depends(get_current_user)):
    oid = to_objid(trip_id)
    trip = trips_col.find_one({"_id": oid, "user_id": user["_id"]})
    if not trip:
        raise HTTPException(404, "Trip not found")
    # ensure lists
    trip["participants"] = trip.get("participants", [])
    trip["expenses"] = trip.get("expenses", [])
    return serialize_trip(trip)


@app.post("/trips/{trip_id}/participant")
def add_participant(trip_id: str, payload: ParticipantCreate = Body(...), user=Depends(get_current_user)):
    oid = to_objid(trip_id)
    trip = trips_col.find_one({"_id": oid, "user_id": user["_id"]})
    if not trip:
        raise HTTPException(404, "Trip not found")

    name = payload.name.strip()
    # avoid duplicate names (case-insensitive)
    existing = [p for p in trip.get("participants", []) if p.get("name", "").lower() == name.lower()]
    if not existing:
        trips_col.update_one({"_id": oid}, {"$push": {"participants": {"name": name}}})
    return {"ok": True, "name": name}


@app.post("/trips/{trip_id}/expense")
def add_trip_expense(trip_id: str, payload: TripExpenseCreate = Body(...), user=Depends(get_current_user)):
    oid = to_objid(trip_id)
    trip = trips_col.find_one({"_id": oid, "user_id": user["_id"]})
    if not trip:
        raise HTTPException(404, "Trip not found")

    # ensure participants contains the payer (auto-add)
    payer = payload.paid_by.strip()
    if payer and payer.lower() not in [p.get("name", "").lower() for p in trip.get("participants", [])]:
        trips_col.update_one({"_id": oid}, {"$push": {"participants": {"name": payer}}})

    created_at_dt = parse_iso_datetime(payload.created_at) if payload.created_at else datetime.utcnow()
    ex = {
        "paid_by": payer,
        "amount": float(payload.amount),
        "description": payload.description or "",
        "created_at": created_at_dt,
    }
    trips_col.update_one({"_id": oid}, {"$push": {"expenses": ex}})
    return {"ok": True, "expense": ex}


def calculate_trip_settlement(trip: dict) -> dict:
    """
    Calculate settlement for a trip (helper function, no DB/auth).
    Returns:
      {
        total: float,
        per_person: float,
        balances: { name: net },  # positive -> should receive, negative -> owes
        lines: [ "A pays ₹X to B", ... ]
      }
    """

    participants = [p.get("name") if isinstance(p, dict) else p for p in trip.get("participants", [])]
    expenses = trip.get("expenses", []) or []

    # If no explicit participants, gather from expenses
    if not participants:
        names = set()
        for e in expenses:
            if e.get("paid_by"):
                names.add(e["paid_by"])
        participants = sorted(list(names))

    n = len(participants)
    if n == 0:
        return {"total": 0.0, "per_person": 0.0, "balances": {}, "lines": []}

    # compute total and per-person
    total = sum((float(e.get("amount") or 0) for e in expenses))
    per_person = total / n if n else 0.0

    # compute net paid per person
    paid_map: Dict[str, float] = {name: 0.0 for name in participants}
    for e in expenses:
        payer = e.get("paid_by")
        amt = float(e.get("amount") or 0)
        if payer not in paid_map:
            # include extra payers not in participants list
            paid_map[payer] = 0.0
            participants.append(payer)
            n = len(participants)
            per_person = total / n if n else 0.0
        paid_map[payer] = paid_map.get(payer, 0.0) + amt

    # balances = paid - share (positive => should receive)
    balances: Dict[str, float] = {}
    for name in participants:
        paid = paid_map.get(name, 0.0)
        balances[name] = round(paid - per_person, 2)

    # Prepare settlement lines using greedy matching
    creditors = [{"name": k, "amount": round(v, 2)} for k, v in balances.items() if v > 0.0]
    debtors = [{"name": k, "amount": round(-v, 2)} for k, v in balances.items() if v < 0.0]

    # Sort creditors desc by amount, debtors desc by amount
    creditors.sort(key=lambda x: x["amount"], reverse=True)
    debtors.sort(key=lambda x: x["amount"], reverse=True)

    lines: List[str] = []
    ci = 0
    di = 0
    # Greedy: largest debtor pays largest creditor
    while di < len(debtors) and ci < len(creditors):
        debt = debtors[di]
        cred = creditors[ci]
        pay = min(debt["amount"], cred["amount"])
        # Round to 2 decimals and to integer rupees if close
        pay_val = round(pay, 2)
        # Format rupee value without trailing .00 when integer
        def fmt_amt(x):
            try:
                if math.isclose(x, round(x)):
                    return str(int(round(x)))
                return f"{x:.2f}"
            except (TypeError, ValueError):
                return str(x)
        lines.append(f"{debt['name']} pays ₹{fmt_amt(pay_val)} to {cred['name']}")
        debtors[di]["amount"] = round(debt["amount"] - pay, 2)
        creditors[ci]["amount"] = round(cred["amount"] - pay, 2)
        if math.isclose(debtors[di]["amount"], 0.0):
            di += 1
        if math.isclose(creditors[ci]["amount"], 0.0):
            ci += 1

    return {
        "total": round(total, 2),
        "per_person": round(per_person, 2),
        "balances": balances,
        "lines": lines,
    }


@app.get("/trips/{trip_id}/settlement")
def trip_settlement(trip_id: str, user=Depends(get_current_user)):
    """Get settlement calculation for a trip."""
    oid = to_objid(trip_id)
    trip = trips_col.find_one({"_id": oid, "user_id": user["_id"]})
    if not trip:
        raise HTTPException(404, "Trip not found")
    return calculate_trip_settlement(trip)




@app.get("/trips/{trip_id}/export")
def export_trip(trip_id: str, token: str = None):
    """
    Export trip as Excel with 4 sheets:
    1. Trip Info (name, dates, participants)
    2. Expenses (paid_by, amount, description)
    3. Balances (person, balance)
    4. Who Pays Whom (from, to, amount)
    Accepts token as query parameter: ?token=YOUR_JWT_TOKEN
    """
    user = None
    
    if token:
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            user = users_col.find_one({"_id": ObjectId(payload["sub"])})
            if not user:
                raise HTTPException(401, "User not found")
        except JWTError:
            raise HTTPException(401, "Invalid token")
    else:
        raise HTTPException(401, "Missing token. Use ?token=YOUR_JWT_TOKEN")
    
    oid = to_objid(trip_id)
    trip = trips_col.find_one({"_id": oid, "user_id": user["_id"]})
    if not trip:
        raise HTTPException(404, "Trip not found")

    settlement = calculate_trip_settlement(trip)
    trip_name = trip.get("trip_name", "Trip")

    # Create workbook
    wb = Workbook()

    # Sheet 1: Trip Info
    ws = wb.active
    ws.title = "Trip Info"
    ws.append(["Field", "Value"])
    ws.append(["Trip Name", trip_name])
    ws.append(["Start Date", str(trip.get("start_date", ""))])
    ws.append(["End Date", str(trip.get("end_date", ""))])
    
    participants = [
        p.get("name", "Unknown") if isinstance(p, dict) else str(p)
        for p in trip.get("participants", [])
    ]
    ws.append(["Participants", ", ".join(participants)])

    # Sheet 2: Expenses
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Paid By", "Amount", "Description", "Date"])
    for e in trip.get("expenses", []):
        ws2.append([
            e.get("paid_by", ""),
            e.get("amount", 0),
            e.get("description", ""),
            str(e.get("created_at", "")),
        ])

    # Sheet 3: Balances
    ws3 = wb.create_sheet("Balances")
    ws3.append(["Person", "Balance (₹)"])
    for person, bal in settlement["balances"].items():
        ws3.append([person, bal])

    # Sheet 4: Who Pays Whom (using lines)
    ws4 = wb.create_sheet("Who Pays Whom")
    ws4.append(["Payment"])
    for line in settlement.get("lines", []):
        ws4.append([line])

    # Add summary to sheet 1
    ws.append(["", ""])
    ws.append(["Total Expenses", settlement["total"]])
    ws.append(["Per Person", settlement["per_person"]])

    # Save to bytes
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trip_{trip_name}.xlsx"}
    )



# ================= RUN =================
if __name__ == "__main__":
    logging.info(JWT_REFRESH_SECRET,EMAIL_FROM)
    print(EMAIL_FROM)
    import uvicorn
    uvicorn.run("main:app", port=8000)
