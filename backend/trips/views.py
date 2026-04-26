from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from openpyxl import Workbook
import io
import tempfile
import os
from .models import Trip, Participant, TripExpense
from .serializers import (
    TripSerializer, TripCreateSerializer,
    TripExpenseCreateSerializer, TripExpenseSerializer
)

class TripListCreateView(generics.ListCreateAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Trip.objects.filter(user=self.request.user)
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TripCreateSerializer
        return TripSerializer
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class TripDetailView(generics.RetrieveDestroyAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Trip.objects.filter(user=self.request.user)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_participant(request, pk):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)

    name = (request.data.get('name') or '').strip()
    if not name:
        return Response({'error': 'Name required'}, status=status.HTTP_400_BAD_REQUEST)

    # Avoid duplicates (case-insensitive)
    existing = trip.participants.filter(name__iexact=name).first()
    if existing:
        return Response({'ok': True, 'name': existing.name, 'id': existing.id})

    participant = trip.participants.create(name=name)
    return Response({'ok': True, 'name': participant.name, 'id': participant.id})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_expense(request, pk):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)

    # Pass trip in serializer context for validation if needed
    serializer = TripExpenseCreateSerializer(data=request.data, context={'trip': trip})
    if serializer.is_valid():
        expense = serializer.save(trip=trip)
        payer = (expense.paid_by or '').strip()
        if payer:
            if not trip.participants.filter(name__iexact=payer).exists():
                trip.participants.create(name=payer)
        return Response({'ok': True, 'expense': TripExpenseSerializer(expense).data})
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_expense(request, pk, expense_id):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)
    try:
        expense = trip.expenses.get(pk=expense_id)
    except TripExpense.DoesNotExist:
        return Response({'error': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = TripExpenseCreateSerializer(expense, data=request.data, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        # Auto-add payer as participant if not exists
        payer = (updated.paid_by or '').strip()
        if payer and not trip.participants.filter(name__iexact=payer).exists():
            trip.participants.create(name=payer)
        return Response({'ok': True, 'expense': TripExpenseSerializer(updated).data})
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_expense(request, pk, expense_id):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)
    try:
        expense = trip.expenses.get(pk=expense_id)
        expense.delete()
        return Response({'ok': True})
    except TripExpense.DoesNotExist:
        return Response({'error': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_participant(request, pk, participant_id):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)
    try:
        participant = trip.participants.get(pk=participant_id)
        participant.delete()
        return Response({'ok': True})
    except Participant.DoesNotExist:
        return Response({'error': 'Participant not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def trip_settlement(request, pk):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)
    
    participants = [p.name for p in trip.participants.all()] if trip.participants.exists() else []
    expenses = trip.expenses.all()
    
    # Gather all unique names from expenses if no participants
    if not participants:
        names = {e.paid_by for e in expenses}
        participants = list(names)
    
    n = len(participants)
    if n == 0:
        return Response({'total': 0, 'per_person': 0, 'balances': {}, 'lines': []})
    
    total = sum(e.amount for e in expenses)
    per_person_raw = total / n
    per_person = round(per_person_raw, 2)
    remainder = round(total - (per_person * n), 2)
    
    # Calculate balances
    paid = {name: 0 for name in participants}
    for e in expenses:
        if e.paid_by in paid:
            paid[e.paid_by] += e.amount
    
    balances = {name: round(paid[name] - per_person, 2) for name in participants}
    
    # Generate settlement lines
    creditors = [(k, v) for k, v in balances.items() if v > 0]
    debtors = [(k, -v) for k, v in balances.items() if v < 0]
    creditors.sort(key=lambda x: x[1], reverse=True)
    debtors.sort(key=lambda x: x[1], reverse=True)
    
    lines = []
    i, j = 0, 0
    while i < len(debtors) and j < len(creditors):
        debtor_name, debt_amt = debtors[i]
        cred_name, cred_amt = creditors[j]
        pay = min(debt_amt, cred_amt)
        lines.append(f"{debtor_name} pays ₹{pay:.2f} to {cred_name}")
        debtors[i] = (debtor_name, debt_amt - pay)
        creditors[j] = (cred_name, cred_amt - pay)
        if debtors[i][1] <= 0.01: i += 1
        if creditors[j][1] <= 0.01: j += 1
    
    # Determine who should absorb remainder (person who paid most, or first participant)
    remainder_suggestion = max(balances, key=balances.get) if balances else (participants[0] if participants else None)
    
    return Response({
        'total': round(total, 2),
        'per_person': per_person,
        'remainder': remainder,
        'remainder_suggestion': remainder_suggestion,
        'balances': balances,
        'lines': lines
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def trip_export(request, pk):
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)

    wb = Workbook()

    # Trip Info
    ws1 = wb.active
    ws1.title = 'Trip Info'
    ws1.append(['Trip Name', trip.trip_name])
    ws1.append(['Participants', ', '.join(p.name for p in trip.participants.all())])

    # Expenses
    ws2 = wb.create_sheet('Expenses')
    ws2.append(['Paid By', 'Amount', 'Description', 'Date'])
    for e in trip.expenses.all():
        ws2.append([e.paid_by, e.amount, e.description, str(e.created_at)])

    # Save to BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="trip_{trip.trip_name}.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def trip_email_report(request, pk):
    """Send trip summary and expenses as Excel attachment via email."""
    trip = Trip.objects.filter(user=request.user, pk=pk).first()
    if not trip:
        return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)

    # Build Excel in memory
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Trip Info'
    ws1.append(['Trip Name', trip.trip_name])
    ws1.append(['Participants', ', '.join(p.name for p in trip.participants.all())])

    ws2 = wb.create_sheet('Expenses')
    ws2.append(['Paid By', 'Amount', 'Description', 'Date'])
    for e in trip.expenses.all():
        ws2.append([e.paid_by, e.amount, e.description, str(e.created_at)])

    # Write to temp file for attachment
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        email = EmailMultiAlternatives(
            subject=f'Trip Report: {trip.trip_name}',
            body=f'Please find the attached report for your trip "{trip.trip_name}".',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[request.user.email]
        )
        email.attach_file(tmp_path)
        email.send()
        return Response({'ok': True, 'message': 'Email sent successfully.'})
    except Exception as e:
        return Response({'ok': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    finally:
        os.unlink(tmp_path)

